import json, glob, os
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

INPUT_DIR = "/Users/robertskalvitis/Downloads/3dgs_llff"
OUTPUT_XLS = os.path.join(os.path.dirname(__file__), "llff_results.xlsx")

SCENES = ["fern", "flower", "fortress", "horns", "leaves", "orchids", "room", "trex"]
ITERS  = ["ours_7000", "ours_30000"]
METRICS = ["PSNR", "SSIM", "LPIPS"]
N_SEEDS = 5

# ── collect ──────────────────────────────────────────────────────────────────
data = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))

for path in glob.glob(os.path.join(INPUT_DIR, "*_seed*/results.json")):
    run = os.path.basename(os.path.dirname(path))
    scene, seed = run.rsplit("_seed", 1)
    with open(path) as f:
        d = json.load(f)
    for itr in ITERS:
        if itr in d:
            for m in METRICS:
                if m in d[itr]:
                    data[scene][itr][m].append(d[itr][m])

# ── build workbook ────────────────────────────────────────────────────────────
wb = Workbook()

HDR_FILL   = PatternFill("solid", fgColor="1F4E79")
HDR_FONT   = Font(color="FFFFFF", bold=True)
SUB_FILL   = PatternFill("solid", fgColor="2E75B6")
SUB_FONT   = Font(color="FFFFFF", bold=True)
AVG_FILL   = PatternFill("solid", fgColor="D6E4F0")
AVG_FONT   = Font(bold=True)
THIN       = Side(style="thin", color="AAAAAA")
BORDER     = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER     = Alignment(horizontal="center", vertical="center")

def hdr(ws, row, col, val, fill=HDR_FILL, font=HDR_FONT, align=CENTER):
    c = ws.cell(row=row, column=col, value=val)
    c.fill, c.font, c.alignment, c.border = fill, font, align, BORDER
    return c

def cell(ws, row, col, val, fill=None, font=None, fmt=None, align=CENTER):
    c = ws.cell(row=row, column=col, value=val)
    if fill: c.fill = fill
    if font: c.font = font
    if fmt:  c.number_format = fmt
    c.alignment, c.border = align, BORDER
    return c

for itr in ITERS:
    label = "7k" if itr == "ours_7000" else "30k"
    ws = wb.create_sheet(title=f"{label} iterations")

    # ── header row 1: metric groups ──
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    hdr(ws, 1, 1, "Scene")
    for mi, m in enumerate(METRICS):
        col = 2 + mi * (N_SEEDS + 1)
        ws.merge_cells(start_row=1, start_column=col,
                       end_row=1,   end_column=col + N_SEEDS)
        hdr(ws, 1, col, m)

    # ── header row 2: seed cols + mean ──
    for mi, m in enumerate(METRICS):
        base = 2 + mi * (N_SEEDS + 1)
        for s in range(N_SEEDS):
            hdr(ws, 2, base + s, f"seed {s}", fill=SUB_FILL, font=SUB_FONT)
        hdr(ws, 2, base + N_SEEDS, "Mean", fill=SUB_FILL, font=SUB_FONT)

    # ── data rows ──
    for ri, scene in enumerate(SCENES):
        row = 3 + ri
        cell(ws, row, 1, scene, font=Font(bold=True),
             align=Alignment(horizontal="left", vertical="center"))

        for mi, m in enumerate(METRICS):
            base  = 2 + mi * (N_SEEDS + 1)
            vals  = data[scene][itr][m]
            fmt   = "0.0000" if m in ("SSIM", "LPIPS") else "0.00"
            # individual seeds (sorted by seed index via filename order)
            seeds_path = sorted(
                glob.glob(os.path.join(INPUT_DIR, f"{scene}_seed*/results.json")))
            seed_vals = []
            for sp in seeds_path:
                with open(sp) as f:
                    sd = json.load(f)
                v = sd.get(itr, {}).get(m)
                if v is not None:
                    seed_vals.append(v)
            for s, v in enumerate(seed_vals):
                cell(ws, row, base + s, round(v, 6), fmt=fmt)
            mean = sum(seed_vals) / len(seed_vals) if seed_vals else None
            cell(ws, row, base + N_SEEDS, round(mean, 6) if mean else None,
                 fill=AVG_FILL, font=AVG_FONT, fmt=fmt)

    # ── overall mean row ──
    row = 3 + len(SCENES)
    cell(ws, row, 1, "Mean", fill=AVG_FILL, font=AVG_FONT,
         align=Alignment(horizontal="left", vertical="center"))
    for mi, m in enumerate(METRICS):
        base = 2 + mi * (N_SEEDS + 1)
        # per-seed column means
        for s in range(N_SEEDS):
            col = base + s
            scene_vals = []
            for scene in SCENES:
                seeds_path = sorted(
                    glob.glob(os.path.join(INPUT_DIR, f"{scene}_seed{s}/results.json")))
                for sp in seeds_path:
                    with open(sp) as f:
                        sd = json.load(f)
                    v = sd.get(itr, {}).get(m)
                    if v is not None:
                        scene_vals.append(v)
            mean = sum(scene_vals) / len(scene_vals) if scene_vals else None
            fmt = "0.0000" if m in ("SSIM", "LPIPS") else "0.00"
            cell(ws, row, col, round(mean, 6) if mean else None,
                 fill=AVG_FILL, font=AVG_FONT, fmt=fmt)
        # overall mean of means
        all_vals = [data[sc][itr][m] for sc in SCENES]
        flat = [v for lst in all_vals for v in lst]
        grand = sum(flat) / len(flat) if flat else None
        fmt = "0.0000" if m in ("SSIM", "LPIPS") else "0.00"
        cell(ws, row, base + N_SEEDS, round(grand, 6) if grand else None,
             fill=PatternFill("solid", fgColor="BDD7EE"), font=Font(bold=True), fmt=fmt)

    # ── column widths ──
    ws.column_dimensions["A"].width = 12
    total_cols = 1 + len(METRICS) * (N_SEEDS + 1)
    for c in range(2, total_cols + 1):
        ws.column_dimensions[get_column_letter(c)].width = 10
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 18
    ws.freeze_panes = "B3"

# remove default sheet
if "Sheet" in wb.sheetnames:
    del wb["Sheet"]

wb.save(OUTPUT_XLS)
print(f"Saved: {OUTPUT_XLS}")
